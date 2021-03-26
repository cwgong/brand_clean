import io
import random
import re
from pyhive import hive
import traceback
import openpyxl


class PddCat3BrandRegFileTool(object):
    def __init__(self, cat1_en_name):
        self.CHECK_BASE_FOLDER = "./check_sample_data"
        self.BASE_FOLDER = "%s_config" % (cat1_en_name)
        self.BRAND_FOLDER = "%s/%s_brand_info" % (self.BASE_FOLDER, cat1_en_name)
        self.FOCUS_BRAND_FILE = "%s/%s_focus_brand.txt" % (self.BRAND_FOLDER, cat1_en_name)
        self.BRAND_ORI_FILE = "%s/%s_brand_ori.txt" % (self.BRAND_FOLDER, cat1_en_name)
        self.CAT3_SAMPLE_FILE = "%s/%s_cat3_check_sample.xlsx" % (self.CHECK_BASE_FOLDER, cat1_en_name)
        self.RANDOM_SAMPLE_FILE = "%s/%s_random_check_sample.xlsx" % (self.CHECK_BASE_FOLDER, cat1_en_name)
        self.TOP_GMV_SKU_FILE = "%s/%s_sku_10w.xlsx" % (self.CHECK_BASE_FOLDER, cat1_en_name)
        self.BRAND_SAMPLE_FILE = "%s/%s_focus_brand_check_sample.xlsx" % (self.CHECK_BASE_FOLDER, cat1_en_name)
        self.BRAND_CAT3_RECALL_FILE = "%s/%s_brand_cat3_recall.txt" % (self.BRAND_FOLDER, cat1_en_name)



def get_new_dt(cur,data_table,cat1_name):
    try:
        sql_str = "SELECT max(dt) from %s where cat1_name='%s'" %(data_table,cat1_name)
        cur.execute(sql_str)
        data = cur.fetchall()
        new_dt = data[0][0]
        return new_dt
    except Exception as e:
        print(traceback.format_exc())

def get_table_count(cur,data_table,cat1_name):
    new_dt = get_new_dt(cur,data_table,cat1_name)
    try:
        sql_str = "SELECT count(*) from %s where cat1_name='%s' and dt='%s'" %(data_table,cat1_name,new_dt)
        cur.execute(sql_str)
        data = cur.fetchall()
        table_count = data[0][0]
        return table_count
    except Exception as e:
        print(traceback.format_exc())

def writeExcel(path, value, sheet):
    '''
    :param sheet:sheet的名称
    :param path:文件的名字和路径
    :param value1: 写入的数据
    :return:
    '''
    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = sheet

    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet1.cell(row=i + 1, column=j + 1, value=str(value[i][j]))

    book.save(path)
    print("写入数据成功！")


def addExcel(path, value, sheet):
    '''
	:param sheet:sheet的名称
    :param path:写入excel的路径
    :param value: 追加的数据
    :return:
    '''
    wb = openpyxl.load_workbook(path)
    wb.create_sheet(sheet)
    ws = wb[sheet]

    for ss in value:
        ws.append(ss)
    wb.save(path)
    print("写入成功")


def get_cat3_sample(brand_file,output_file,cat1_name,table_name):
    '''
    进行三级类下的随机采样
    :param brand_file:
    :param output_file:
    :return:
    '''
    sample_check_list = []
    cat3_dict = {}
    save_value_list = []

    with io.open(brand_file,"r",encoding="utf-8") as f1:
        for line in f1:
            if len(line) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 6:continue
            brand_id,brand_recall_name,brand_name,cat1,cat2,cat3 = line_list
            cat3_dict[cat3] = ""
    print(len(cat3_dict))
    cat3_list = list(cat3_dict.keys())
    if len(cat3_list) < 10:
        cat3_sample_list = cat3_list
    elif len(cat3_list) < 30 and len(cat3_list) >= 10:
        cat3_sample_list = random.sample(cat3_list, 10)  # 从list中随机获取10个元素，作为一个片断返回
    elif len(cat3_list) < 50 and len(cat3_list) >= 30:
        cat3_sample_list = random.sample(cat3_list, 20)  # 从list中随机获取20个元素，作为一个片断返回
    elif len(cat3_list) < 80 and len(cat3_list) >= 50:
        cat3_sample_list = random.sample(cat3_list, 30)  # 从list中随机获取30个元素，作为一个片断返回
    elif len(cat3_list) < 100 and len(cat3_list) >= 80:
        cat3_sample_list = random.sample(cat3_list, 40)  # 从list中随机获取40个元素，作为一个片断返回
    else:
        cat3_sample_list = random.sample(cat3_list, 50)  # 从list中随机获取50个元素，作为一个片断返回


    conn = hive.connect(host='172.20.207.6', port=10000, username='supdev')
    cur = conn.cursor()
    new_dt = get_new_dt(cur,table_name,cat1_name)

    cat3_name_tmp = ["'" + sample_item + "'" for sample_item in cat3_sample_list]
    cat3_name_str = '(' + ','.join(cat3_name_tmp) + ')'

    try:
        sql1 = """
    select * 
      from 
      (
        select * ,row_number() over(partition by x.cat3_std_name order by rand())rn
         from 
            (
              select * from dwd.dwd_pdd_cat3_brand_reg a
              where a.dt = '%s' and a.cat1_name='%s' and a.cat3_std_name in %s
            )x
      )d
    where rn <= 150""" % (new_dt, cat1_name, cat3_name_str)
        cur.execute(sql1)
        data = cur.fetchall()
        sample_check_list.append(data)
    except Exception as e:
        print(traceback.format_exc())

    for epoch in sample_check_list:
        for item in epoch:
            save_value_list.append(item)

    save_value_list.insert(0, ['sku_id', 'title', 'brand_std_id', 'brand_std_name', 'match_brand_name','cat1_id', 'cat1_name', 'cat2_id',
                               'cat2_name', \
                               'cat3_id', 'cat3_name'])

    writeExcel(output_file,save_value_list,'三级类品牌采样')


def get_random_sample(output_file,cat1_name,data_table):
    '''
    所有商品下的随机采样
    :param output_file:
    :return:
    '''

    conn = hive.connect(host='172.20.207.6', port=10000, username='supdev')
    cur = conn.cursor()

    new_dt = get_new_dt(cur,data_table,cat1_name)

    sku_count = get_table_count(cur,data_table,cat1_name)

    no_dict = {}
    sample_check_list = []
    save_value_list = []

    while True:
        tmp = random.randint(0, sku_count)
        if tmp not in no_dict:
            no_dict[tmp] = ''
        if len(no_dict) >= 3000:
            break
    r_lst = []
    for k, v in no_dict.items():
        r_lst.append(str(k))

    r_lst_tmp = ["'" + str(item) + "'" for item in r_lst]
    where_cond = "(" + ", ".join(r_lst_tmp) + ")"

    try:
        sql_str = """
        select
        x.sku_id, x.title, x.brand_std_id, x.brand_std_name, x.cat1_std_id, x.cat1_std_name, x.cat2_std_id, x.cat2_std_name, x.cat3_std_id, x.cat3_std_name
        from
        (
            select row_number()
        over(partition
        by
        1) as rw_no, sku_id, title, brand_std_id, brand_std_name, cat1_std_id, cat1_std_name, cat2_std_id, cat2_std_name, cat3_std_id, cat3_std_name
        from dwd.dwd_pdd_cat3_brand_reg
            where
        cat1_name = '%s' and dt = '%s'
        ) x
        where rw_no in %s""" %(cat1_name,new_dt,where_cond)
        cur.execute(sql_str)
        data = cur.fetchall()
        sample_check_list.append(data)
    except Exception as e:
        print(traceback.format_exc())

    for epoch in sample_check_list:
        for item in epoch:
            save_value_list.append(item)

    save_value_list.insert(0,['sku_id','title','brand_std_id','brand_std_name','cat1_id','cat1_name','cat2_id','cat2_name',\
                              'cat3_id','cat3_name'])

    writeExcel(output_file, save_value_list, '全局随机采样')


def get_seed_brand_info(input_file):
    seed_brand_dict = {}
    with open(input_file, "r", encoding="utf-8") as f1:
        for line in f1:
            line = line.strip()
            if line == "": continue
            brand_id,brand_name = line.split("\t")
            seed_brand_dict[brand_id] = brand_name
    return seed_brand_dict


def get_brand_topgmv_sample(focus_brand_file,output_check_file,cat1_name,data_table):
    '''
    重点品牌维度下进行topGMV采样
    :param focus_brand_file:
    :param output_check_file:
    :return:
    '''
    sample_check_list = []
    save_value_list = []

    seed_brand_dict = get_seed_brand_info(focus_brand_file)
    seed_brand_list = [(k,v) for k,v in seed_brand_dict.items()]
    sample_brand_list = seed_brand_list[:15]
    extra_brand_list = seed_brand_list[15:]
    sample_brand_list = sample_brand_list + random.sample(extra_brand_list,35)

    conn = hive.connect(host='172.20.207.6', port=10000, username='supdev')
    cur = conn.cursor()
    new_dt = get_new_dt(cur,data_table,cat1_name)

    brand_id_tmp = ["'" + sample_item[0] + "'" for sample_item in sample_brand_list]
    brand_id_str = '(' + ','.join(brand_id_tmp) + ')'

    try:
        sql1 = """
select * 
  from 
  (
    select * ,row_number() over(partition by x.brand_std_id order by x.gmv desc)rn
     from 
        (
          select a.*,c.gmv from %s a
          left join (SELECT sku_id,max(title) title,sum(sale_amount) AS gmv
            FROM dwi.dwi_retailers_online_platform_info
            WHERE platform_type = 'pdd'
            AND dc = 'month'
            group by sku_id) c on c.sku_id = a.sku_id
            where a.dt = '%s' and c.gmv is not null and a.brand_std_id in %s and a.cat1_name='%s'
        )x
  )d
where rn <= 60""" % (data_table,new_dt,brand_id_str,cat1_name)
        cur.execute(sql1)
        data = cur.fetchall()
        sample_check_list.append(data)
    except Exception as e:
        print(traceback.format_exc())

    for epoch in sample_check_list:
        for item in epoch:
            save_value_list.append(item)

    save_value_list.insert(0, ['sku_id', 'title', 'brand_std_id', 'brand_std_name', 'match_type_name','cat1_id', 'cat1_name', 'cat2_id',
                               'cat2_name', \
                               'cat3_id', 'cat3_name'])

    writeExcel(output_check_file,save_value_list,'重点品牌topGMV采样')


def get_sku_10w(output_check_file,cat1_name,data_table):
    sample_check_list = []
    save_value_list = []
    conn = hive.connect(host='172.20.207.6', port=10000, username='supdev')
    cur = conn.cursor()
    new_dt = get_new_dt(cur, data_table, cat1_name)
    try:
        sql_str = """select * from %s a
        left join dwi.dwi_retailers_online_platform_info_pdd_10w
        c on c.sku_id = a.sku_id
        where a.dt = '%s' and c.sku_id is not null and a.cat1_name='%s'""" %(data_table,new_dt,cat1_name)

        cur.execute(sql_str)
        data_tuple = cur.fetchall()
        sample_check_list.append(data_tuple)
    except Exception as e:
        print(traceback.format_exc())

    for epoch in sample_check_list:
        for item in epoch:
            save_value_list.append(item)

    save_value_list.insert(0, ['sku_id', 'title', 'brand_std_id', 'brand_std_name', 'match_type_name', 'cat1_id',
                               'cat1_name', 'cat2_id',
                               'cat2_name', \
                               'cat3_id', 'cat3_name'])
    writeExcel(output_check_file, save_value_list, '10w+商品')


if __name__ == "__main__":

    # #更新后的采样方式（不关联gmv），全局随机采样
    cat1_name = 'fushineiyi'
    pdd_file_sys = PddCat3BrandRegFileTool(cat1_name)
    data_table = 'dwd.dwd_pdd_cat3_brand_reg'
    get_random_sample(pdd_file_sys.RANDOM_SAMPLE_FILE,cat1_name,data_table)

    # # 重点品牌维度下进行topGMV采样
    # cat1_name = 'shoujipeijian'
    # data_table = 'dwd.dwd_pdd_cat3_brand_reg'
    # pdd_file_sys = PddCat3BrandRegFileTool(cat1_name)
    # get_brand_topgmv_sample(pdd_file_sys.FOCUS_BRAND_FILE,pdd_file_sys.BRAND_SAMPLE_FILE,cat1_name,data_table)

    #随机选择三级类并随机进行采样
    # cat1_name = 'shuma'
    # table_name = 'dwd.dwd_pdd_cat3_brand_reg'
    # pdd_cat3_sys = PddCat3BrandRegFileTool(cat1_name)
    # get_cat3_sample(pdd_cat3_sys.BRAND_CAT3_RECALL_FILE,pdd_cat3_sys.CAT3_SAMPLE_FILE,cat1_name,table_name)

    # 取到所有10w+数据
    # cat1_name = 'shuma'
    # pdd_file_sys = PddCat3BrandRegFileTool(cat1_name)
    # data_table = 'dwd.dwd_pdd_cat3_brand_reg'
    # get_sku_10w(pdd_file_sys.TOP_GMV_SKU_FILE, cat1_name, data_table)




