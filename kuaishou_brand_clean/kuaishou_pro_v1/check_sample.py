import io
import random
import re
from pyhive import hive
import traceback
import openpyxl


class PddCat3BrandRegFileTool(object):
    def __init__(self, cat1_en_name):
        self.CHECK_BASE_FOLDER = "./check_sample_data"
        self.CAT3_SAMPLE_FILE = "%s/%s_cat3_check_sample.xlsx" % (self.CHECK_BASE_FOLDER, cat1_en_name)
        self.RANDOM_SAMPLE_FILE = "%s/%s_random_check_sample.xlsx" % (self.CHECK_BASE_FOLDER, cat1_en_name)
        self.TOP_GMV_SKU_FILE = "%s/%s_sku_10w.xlsx" % (self.CHECK_BASE_FOLDER, cat1_en_name)
        self.BRAND_SAMPLE_FILE = "%s/%s_focus_brand_check_sample.xlsx" % (self.CHECK_BASE_FOLDER, cat1_en_name)


def get_new_dt(cur,data_table):
    try:
        sql_str = "SELECT max(dt) from %s" %(data_table)
        cur.execute(sql_str)
        data = cur.fetchall()
        new_dt = data[0][0]
        return new_dt
    except Exception as e:
        print(traceback.format_exc())

def get_table_count(cur,data_table):
    new_dt = get_new_dt(cur,data_table)
    try:
        sql_str = "SELECT count(*) from %s where dt='%s'" %(data_table,new_dt)
        cur.execute(sql_str)
        data = cur.fetchall()
        table_count = data[0][0]
        return table_count
    except Exception as e:
        print(traceback.format_exc())


def writeExcel(path, value, sheet):
    '''
    :param sheet:sheet的名称
    :param path:文件的名字和路径
    :param value1: 写入的数据
    :return:
    '''
    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = sheet

    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet1.cell(row=i + 1, column=j + 1, value=str(value[i][j]))

    book.save(path)
    print("写入数据成功！")


def addExcel(path, value, sheet):
    '''
	:param sheet:sheet的名称
    :param path:写入excel的路径
    :param value: 追加的数据
    :return:
    '''
    wb = openpyxl.load_workbook(path)
    wb.create_sheet(sheet)
    ws = wb[sheet]

    for ss in value:
        ws.append(ss)
    wb.save(path)
    print("写入成功")


def get_random_sample(output_file,data_table):
    '''
    所有商品下的随机采样
    :param output_file:
    :return:
    '''

    conn = hive.connect(host='172.20.207.6', port=10000, username='supdev')
    cur = conn.cursor()

    new_dt = get_new_dt(cur,data_table)

    sku_count = get_table_count(cur,data_table)

    no_dict = {}
    sample_check_list = []
    save_value_list = []

    while True:
        tmp = random.randint(0, sku_count)
        if tmp not in no_dict:
            no_dict[tmp] = ''
        if len(no_dict) >= 3000:
            break
    r_lst = []
    for k, v in no_dict.items():
        r_lst.append(str(k))

    r_lst_tmp = ["'" + str(item) + "'" for item in r_lst]
    where_cond = "(" + ", ".join(r_lst_tmp) + ")"

    try:
        sql_str = """
        select
        x.product_id, x.product_name, x.brand_id_std, x.brand_name_std, 
        x.category1_id_std, x.category1_std
        from
        (
            select row_number()
        over(partition
        by
        1) as rw_no, product_id, product_name , brand_id_std , brand_name_std , 
        category1_id_std  , category1_std  
        from %s
        where
        dt = '%s'
        ) x
        where rw_no in %s""" %(data_table,new_dt,where_cond)
        cur.execute(sql_str)
        data = cur.fetchall()
        sample_check_list.append(data)
    except Exception as e:
        print(traceback.format_exc())

    for epoch in sample_check_list:
        for item in epoch:
            save_value_list.append(item)

    save_value_list.insert(0,['sku_id','title','brand_std_id','brand_std_name','cat1_id','cat1_name'])

    writeExcel(output_file, save_value_list, '全局随机采样')

if __name__ == "__main__":
    # 更新后的采样方式（不关联gmv），全局随机采样
    cat_name = 'kuaishou'
    pdd_file_sys = PddCat3BrandRegFileTool(cat_name)
    data_table = 'dim.dim_kuaishou_brand_std_wy'
    get_random_sample(pdd_file_sys.RANDOM_SAMPLE_FILE,data_table)